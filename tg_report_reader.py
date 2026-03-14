import os
import re
import math
import pandas as pd

from datetime import datetime, timedelta, timezone, time
from telethon import TelegramClient
from config import API_ID, API_HASH, SESSION_NAME, TIMEZONE_OFFSET

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


MY_TZ = timezone(timedelta(hours=TIMEZONE_OFFSET))

# 只抓开头为 BCSG + 数字
GROUP_PATTERN = re.compile(r"^(BCSG\d+)", re.IGNORECASE)


# =========================
# 基础工具
# =========================
def extract_bcsg_code(name: str):
    if not name:
        return None
    m = GROUP_PATTERN.match(name.strip())
    return m.group(1).upper() if m else None


def safe_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(",", "").replace("，", "").replace(" ", "")
    try:
        return float(text)
    except Exception:
        return None


def fmt_amount(x):
    if x is None:
        return "-"
    x = float(x)
    if x.is_integer():
        return f"{int(x):,}"
    return f"{x:,.2f}"


def to_excel_number(x):
    if x is None:
        return None
    return float(x)


def now_local():
    return datetime.now(MY_TZ)


def normalize_text(text: str):
    if not text:
        return ""
    return text.replace("\u00A0", " ").replace("：", ":").replace("（", "(").replace("）", ")")


def extract_first_number(pattern, text, flags=re.IGNORECASE):
    if not text:
        return None
    m = re.search(pattern, text, flags)
    if not m:
        return None
    return safe_float(m.group(1))


def extract_first_int(pattern, text, flags=re.IGNORECASE):
    if not text:
        return None
    m = re.search(pattern, text, flags)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


# =========================
# 报表解析
# =========================
def parse_today_report(text: str):
    """
    今日实时数据识别：
    已入账 (x笔)
    已出账 (x笔)
    总已入账
    总已出账
    总入款额
    """
    if not text:
        return None

    text = normalize_text(text)

    in_count = extract_first_int(r"已入账\s*\((\d+)\s*笔\)", text)
    out_count = extract_first_int(r"已出账\s*\((\d+)\s*笔\)", text)

    total_in = extract_first_number(r"总已入账\s*:\s*([-\d,\.]+)", text)
    total_out = extract_first_number(r"总已出账\s*:\s*([-\d,\.]+)", text)

    # 兼容不同写法
    if total_in is None:
        total_in = extract_first_number(r"总入账\s*:\s*([-\d,\.]+)", text)

    # 如果没有 总已入账 / 总入账，就用 总入款额
    fallback_total_in = None
    if total_in is None:
        fallback_total_in = extract_first_number(r"总入款额\s*:\s*([-\d,\.]+)", text)
        total_in = fallback_total_in

    if total_out is None:
        total_out = extract_first_number(r"总出账\s*:\s*([-\d,\.]+)", text)

    # 只要 这3个字段 任何一个有值，就算有效报表
    if total_in is None and total_out is None:
        return None

    # 没有就补 0，方便后面统一显示
    if total_in is None:
        total_in = 0.0

    if total_out is None:
        total_out = 0.0

    # 过滤完全 0 / 0
    if total_in == 0 and total_out == 0:
        return None

    return {
        "in_count": in_count if in_count is not None else 0,
        "out_count": out_count if out_count is not None else 0,
        "total_in": total_in,
        "total_out": total_out,
        "p_hold": None,
        "unpaid": None,
    }


def parse_yesterday_report(text: str):
    """
    昨日导出Excel识别：
    已入账 (x笔)
    已下发 (x笔)
    总入款额
    应下发
    已下发
    未下发
    """
    if not text:
        return None

    text = normalize_text(text)

    in_count = extract_first_int(r"已入账\s*\((\d+)\s*笔\)", text)
    sent_count = extract_first_int(r"已下发\s*\((\d+)\s*笔\)", text)

    total_deposit = extract_first_number(r"总入款额\s*:\s*([-\d,\.]+)", text)
    if total_deposit is None:
        total_deposit = extract_first_number(r"总已入账\s*:\s*([-\d,\.]+)", text)

    should_send = extract_first_number(r"应下发\s*:\s*([-\d,\.]+)", text)

    # 这里为了避免和“已下发(x笔)”冲突，强制带冒号
    sent_amount = extract_first_number(r"已下发\s*:\s*([-\d,\.]+)", text)
    unpaid = extract_first_number(r"未下发\s*:\s*([-\d,\.]+)", text)

    # 昨日报表至少要有 总入款额 + 未下发
    if total_deposit is None or unpaid is None:
        return None

    # 过滤完全 0/0
    if total_deposit == 0 and unpaid == 0:
        return None

    return {
        "in_count": in_count if in_count is not None else 0,
        "out_count": sent_count if sent_count is not None else 0,
        "total_in": total_deposit,
        "total_out": sent_amount if sent_amount is not None else 0.0,
        "p_hold": should_send,
        "unpaid": unpaid,
    }


# =========================
# Telegram 读取
# =========================
async def get_client():
    client = TelegramClient(SESSION_NAME, API_ID, API_HASH)
    await client.start()
    return client


async def get_bcsg_groups(client):
    dialogs = await client.get_dialogs()
    groups = []

    for dialog in dialogs:
        code = extract_bcsg_code(dialog.name)
        if code:
            groups.append({
                "dialog": dialog,
                "code": code
            })

    def sort_key(item):
        m = re.search(r"\d+", item["code"])
        return int(m.group()) if m else 999999

    groups.sort(key=sort_key)
    return groups


async def find_today_latest_valid_report(client, dialog, limit=3000):
    today = now_local().date()

    async for msg in client.iter_messages(dialog.id, limit=limit):
        if not msg.message:
            continue

        dt = msg.date.astimezone(MY_TZ)

        if dt.date() != today:
            continue

        parsed = parse_today_report(msg.message)
        if parsed:
            # 因为 iter_messages 默认从新到旧
            # 第一条符合条件的就是今天最新有效报表
            return dt, parsed

    return None


async def find_yesterday_before_noon_latest_valid_report(client, dialog, limit=3000):
    yesterday = (now_local() - timedelta(days=1)).date()
    cutoff = time(12, 0)

    async for msg in client.iter_messages(dialog.id, limit=limit):
        if not msg.message:
            continue

        dt = msg.date.astimezone(MY_TZ)

        # 只抓昨天
        if dt.date() != yesterday:
            continue

        # 只抓 12:00 前
        if dt.time() >= cutoff:
            continue

        parsed = parse_yesterday_report(msg.message)
        if parsed:
            # 从新到旧扫，所以第一条就是“昨天12点前最后一条有效报表”
            return dt, parsed

    return None


# =========================
# 资料汇总
# =========================
def build_summary_dict(results):
    total_in_sum = 0.0
    total_out_sum = 0.0
    total_in_count = 0
    total_out_count = 0
    total_p_hold = 0.0
    total_unpaid = 0.0
    p_hold_count = 0
    unpaid_count = 0

    for item in results:
        total_in_sum += item["total_in"] or 0.0
        total_out_sum += item["total_out"] or 0.0
        total_in_count += item["in_count"] or 0
        total_out_count += item["out_count"] or 0

        if item["p_hold"] is not None:
            total_p_hold += item["p_hold"]
            p_hold_count += 1

        if item["unpaid"] is not None:
            total_unpaid += item["unpaid"]
            unpaid_count += 1

    return {
        "total_in_sum": total_in_sum,
        "total_out_sum": total_out_sum,
        "total_in_count": total_in_count,
        "total_out_count": total_out_count,
        "total_p_hold": total_p_hold if p_hold_count > 0 else None,
        "total_unpaid": total_unpaid if unpaid_count > 0 else None,
    }


async def collect_today_reports():
    client = await get_client()

    try:
        groups = await get_bcsg_groups(client)

        results = []
        missing = []

        for g in groups:
            report = await find_today_latest_valid_report(client, g["dialog"])

            if not report:
                missing.append(g["code"])
                continue

            dt, parsed = report

            results.append({
                "code": g["code"],
                "date": dt.strftime("%d-%m-%Y"),
                "time": dt.strftime("%H:%M"),
                "total_in": parsed["total_in"],
                "in_count": parsed["in_count"],
                "total_out": parsed["total_out"],
                "out_count": parsed["out_count"],
                "p_hold": parsed["p_hold"],
                "unpaid": parsed["unpaid"],
            })

        return {
            "date": now_local().strftime("%d-%m-%Y"),
            "updated_at": now_local().strftime("%H:%M"),
            "results": results,
            "missing": missing,
            "summary": build_summary_dict(results),
            "group_count": len(groups),
        }

    finally:
        await client.disconnect()


async def collect_yesterday_reports_before_noon():
    client = await get_client()

    try:
        groups = await get_bcsg_groups(client)
        results = []
        missing = []

        for g in groups:
            report = await find_yesterday_before_noon_latest_valid_report(client, g["dialog"])

            if not report:
                missing.append(g["code"])
                continue

            dt, parsed = report

            results.append({
                "code": g["code"],
                "date": dt.strftime("%d-%m-%Y"),
                "time": dt.strftime("%H:%M"),
                "total_in": parsed["total_in"],
                "in_count": parsed["in_count"],
                "total_out": parsed["total_out"],
                "out_count": parsed["out_count"],
                "p_hold": parsed["p_hold"],
                "unpaid": parsed["unpaid"],
            })

        yesterday = (now_local() - timedelta(days=1)).strftime("%d-%m-%Y")

        return {
            "date": yesterday,
            "updated_at": now_local().strftime("%H:%M"),
            "results": results,
            "missing": missing,
            "summary": build_summary_dict(results),
            "group_count": len(groups),
            "rule": "昨天 12:00 前最后一条有效报表",
        }

    finally:
        await client.disconnect()


# =========================
# 文字输出
# =========================
def build_realtime_text(data):
    lines = []
    lines.append("📊 BC Auto Report · 当前实时数据")
    lines.append("")
    lines.append(f"📅 日期：{data['date']}")
    lines.append(f"⏰ 更新时间：{data['updated_at']}")
    lines.append("")

    if not data["results"]:
        lines.append("今天找不到有效报表。")
        return "\n".join(lines)

    for item in data["results"]:
        lines.append(f"🏦 {item['code']}")
        lines.append(f"时间：{item['time']}")
        lines.append(f"总入账：{fmt_amount(item['total_in'])} ({item['in_count']}笔)")
        lines.append(f"总出账：-{fmt_amount(item['total_out'])} ({item['out_count']}笔)")

        if item["p_hold"] is not None:
            lines.append(f"P寄存：{fmt_amount(item['p_hold'])}")
        if item["unpaid"] is not None:
            lines.append(f"未下发：{fmt_amount(item['unpaid'])}")

        lines.append("")

    s = data["summary"]

    lines.append("━━━━━━━━━━━━━━")
    lines.append("")
    lines.append("📈 总结账单")
    lines.append(f"日期：{data['date']}")
    lines.append(f"总入账：{fmt_amount(s['total_in_sum'])} ({s['total_in_count']}笔)")
    lines.append(f"总出账：-{fmt_amount(s['total_out_sum'])} ({s['total_out_count']}笔)")

    if s["total_p_hold"] is not None:
        lines.append(f"P寄存合计：{fmt_amount(s['total_p_hold'])}")
    if s["total_unpaid"] is not None:
        lines.append(f"未下发合计：{fmt_amount(s['total_unpaid'])}")

    return "\n".join(lines)


def build_summary_text(data):
    s = data["summary"]

    lines = [
        "📈 BC Auto Report · 今日汇总",
        "",
        f"📅 日期：{data['date']}",
        f"⏰ 更新时间：{data['updated_at']}",
        "",
        f"总入账：{fmt_amount(s['total_in_sum'])} ({s['total_in_count']}笔)",
        f"总出账：-{fmt_amount(s['total_out_sum'])} ({s['total_out_count']}笔)",
    ]

    if s["total_p_hold"] is not None:
        lines.append(f"P寄存合计：{fmt_amount(s['total_p_hold'])}")
    if s["total_unpaid"] is not None:
        lines.append(f"未下发合计：{fmt_amount(s['total_unpaid'])}")

    lines.extend([
        "",
        f"有效群数：{len(data['results'])}",
        f"缺报表群数：{len(data['missing'])}",
        f"全部BC群数：{data.get('group_count', len(data['results']) + len(data['missing']))}",
    ])

    return "\n".join(lines)


def build_group_check_text(data):
    lines = []
    lines.append("📡 BC Auto Report · 检查所有BC群")
    lines.append("")
    lines.append(f"✅ 有效报表群：{len(data['results'])}")
    lines.append(f"⚠️ 缺报表群：{len(data['missing'])}")
    lines.append(f"🏦 全部BC群：{data.get('group_count', len(data['results']) + len(data['missing']))}")
    lines.append("")

    if data["missing"]:
        lines.append("缺报表群名单：")
        for code in data["missing"]:
            lines.append(f"- {code}")
    else:
        lines.append("今天所有 BCSG 群都有有效报表。")

    return "\n".join(lines)


def build_single_group_text(data, code):
    for item in data["results"]:
        if item["code"] == code:
            lines = [
                "🔍 BC Auto Report · 查询单个BCSG",
                "",
                f"🏦 {item['code']}",
                f"📅 日期：{item['date']}",
                f"⏰ 时间：{item['time']}",
                f"总入账：{fmt_amount(item['total_in'])} ({item['in_count']}笔)",
                f"总出账：-{fmt_amount(item['total_out'])} ({item['out_count']}笔)",
            ]

            if item["p_hold"] is not None:
                lines.append(f"P寄存：{fmt_amount(item['p_hold'])}")
            if item["unpaid"] is not None:
                lines.append(f"未下发：{fmt_amount(item['unpaid'])}")

            return "\n".join(lines)

    return f"{code} 今天没有有效报表。"


def build_system_status_text(today_data):
    lines = [
        "⚙️ BC Auto Report · 系统状态",
        "",
        "状态：运行中",
        "模块：实时数据 / 今日汇总 / 单群查询 / BC群检查 / Excel导出",
        "昨日规则：昨天 12:00 前最后一条有效报表",
        "",
        f"当前时间：{now_local().strftime('%d-%m-%Y %H:%M')}",
        f"识别到BC群总数：{today_data.get('group_count', 0)}",
        f"今日有效报表群：{len(today_data.get('results', []))}",
        f"今日缺报表群：{len(today_data.get('missing', []))}",
    ]
    return "\n".join(lines)


# =========================
# 对外函数
# =========================
async def get_today_realtime_text():
    data = await collect_today_reports()
    return build_realtime_text(data)


async def get_today_summary_text():
    data = await collect_today_reports()
    return build_summary_text(data)


async def get_group_check_text():
    data = await collect_today_reports()
    return build_group_check_text(data)


async def get_single_group_text(code):
    data = await collect_today_reports()
    return build_single_group_text(data, code)


async def get_system_status_text():
    data = await collect_today_reports()
    return build_system_status_text(data)


# =========================
# Excel 美化
# =========================
def auto_fit_columns(ws, min_width=12, max_width=28):
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        width = max(min_width, min(max_len + 2, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_body(cell, is_number=False):
    thin = Side(style="thin", color="D9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(horizontal="center" if not is_number else "right", vertical="center")
    if is_number:
        cell.number_format = '#,##0.00'


def style_title_block(ws, report_title, report_date, updated_at, effective_count, missing_count, extra_rule=None):
    ws["A1"] = report_title
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0F243E")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws["A2"] = f"报表日期：{report_date}"
    ws["A3"] = f"生成时间：{updated_at}"
    ws["A4"] = f"有效群数：{effective_count}"
    ws["A5"] = f"缺报表群数：{missing_count}"

    if extra_rule:
        ws["A6"] = f"规则：{extra_rule}"

    for row in range(1, 7):
        ws.row_dimensions[row].height = 20


def write_summary_block(ws, start_row, summary):
    ws.cell(start_row, 1, "汇总项目")
    ws.cell(start_row, 2, "金额 / 数量")
    style_header(ws.cell(start_row, 1))
    style_header(ws.cell(start_row, 2))

    rows = [
        ("总入账金额", summary.get("total_in_sum")),
        ("总入账笔数", summary.get("total_in_count")),
        ("总出账金额", summary.get("total_out_sum")),
        ("总出账笔数", summary.get("total_out_count")),
        ("P寄存合计", summary.get("total_p_hold")),
        ("未下发合计", summary.get("total_unpaid")),
    ]

    r = start_row + 1
    for label, value in rows:
        ws.cell(r, 1, label)
        ws.cell(r, 2, value if value is not None else "")
        style_body(ws.cell(r, 1), is_number=False)
        style_body(ws.cell(r, 2), is_number=isinstance(value, (int, float)))
        if isinstance(value, (int, float)):
            ws.cell(r, 2).number_format = '#,##0.00'
        r += 1


def format_report_workbook(filename):
    wb = load_workbook(filename)

    # 主Sheet
    ws = wb["报表明细"]
    ws.freeze_panes = "A8"

    # 标题区、表头已写好，这里统一样式
    header_row = 8
    for col in range(1, ws.max_column + 1):
        style_header(ws.cell(header_row, col))

    numeric_cols = {"D", "E", "F", "G", "H", "I"}

    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            col_letter = get_column_letter(col)
            style_body(cell, is_number=(col_letter in numeric_cols))

    auto_fit_columns(ws)

    # 缺报表Sheet
    if "缺报表群" in wb.sheetnames:
        ws2 = wb["缺报表群"]
        ws2.freeze_panes = "A2"
        for col in range(1, ws2.max_column + 1):
            style_header(ws2.cell(1, col))
        for row in range(2, ws2.max_row + 1):
            for col in range(1, ws2.max_column + 1):
                style_body(ws2.cell(row, col), is_number=False)
        auto_fit_columns(ws2)

    wb.save(filename)


def create_report_excel(data, filename, report_title, extra_rule=None):
    rows = []
    for item in data["results"]:
        rows.append({
            "日期": item["date"],
            "时间": item["time"],
            "群名": item["code"],
            "入账笔数": item["in_count"],
            "出账笔数": item["out_count"],
            "总已入账": to_excel_number(item["total_in"]),
            "总已出账": to_excel_number(item["total_out"]),
            "P寄存": to_excel_number(item["p_hold"]),
            "未下发": to_excel_number(item["unpaid"]),
        })

    columns = ["日期", "时间", "群名", "入账笔数", "出账笔数", "总已入账", "总已出账", "P寄存", "未下发"]
    df = pd.DataFrame(rows, columns=columns)

    missing_df = pd.DataFrame({
        "缺报表群名": data["missing"]
    })

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # 先写主明细表，从第8行开始
        df.to_excel(writer, sheet_name="报表明细", index=False, startrow=7)

        ws = writer.book["报表明细"]

        # 标题区
        style_title_block(
            ws=ws,
            report_title=report_title,
            report_date=data["date"],
            updated_at=data["updated_at"],
            effective_count=len(data["results"]),
            missing_count=len(data["missing"]),
            extra_rule=extra_rule,
        )

        # 汇总区
        write_summary_block(ws, start_row=1, summary=data["summary"])

        # 缺报表Sheet
        missing_df.to_excel(writer, sheet_name="缺报表群", index=False)

    format_report_workbook(filename)
    return filename


# =========================
# Excel 导出
# =========================
async def export_today_excel():
    data = await collect_today_reports()
    filename = f"bc_today_report_{data['date']}.xlsx"
    return create_report_excel(
        data=data,
        filename=filename,
        report_title="BC Auto Report - 今日实时数据报表",
        extra_rule="今日最新有效报表"
    )


async def export_yesterday_excel():
    data = await collect_yesterday_reports_before_noon()
    filename = f"bc_yesterday_report_{data['date']}.xlsx"
    return create_report_excel(
        data=data,
        filename=filename,
        report_title="BC Auto Report - 昨日数据报表",
        extra_rule="昨天 12:00 前最后一条有效报表"
    )
